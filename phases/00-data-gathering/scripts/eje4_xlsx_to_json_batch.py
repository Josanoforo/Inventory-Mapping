#!/usr/bin/env python3
"""Pre-procesa catalogos_eje4_canal_descubrimiento.xlsx a JSON individuales
por query + batch_manifest.json para consumo del agente phase0-eje4-discovery.

Spec: agents/codex/phase0-eje4-discovery/README.md sección "Script de
pre-procesamiento". D-169 removió físicamente surfaces gap del xlsx, por lo
que el script solo valida (no filtra).
"""
from __future__ import annotations

import json
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

SHEETS = ("catalogo_1", "catalogo_2", "catalogo_3a", "catalogo_3b")
COLUMNS = (
    "query_id", "catalogo", "tema_semilla", "pattern_id", "query_text",
    "idioma", "region", "surface", "metodo_pago_variable",
    "canal_alternativo", "ventana_temporal", "notes_operador",
)
REQUIRED = ("query_id", "catalogo", "query_text", "idioma", "region", "surface", "ventana_temporal")
SURFACE_ENUM = ("reddit", "blog", "medium", "forum")


def load_rows(xlsx_path: Path, warnings: list[dict]) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if tuple(wb.sheetnames) != SHEETS:
        raise SystemExit(f"Sheet mismatch: expected {SHEETS}, got {tuple(wb.sheetnames)}")
    rows: list[dict] = []
    for sheet_name in SHEETS:
        ws = wb[sheet_name]
        headers = tuple(c.value for c in ws[1])
        if headers != COLUMNS:
            raise SystemExit(f"Schema mismatch in '{sheet_name}': headers = {headers}")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None or not str(row[0]).strip():
                continue
            record = {col: row[i] for i, col in enumerate(COLUMNS)}
            record["_sheet"] = sheet_name
            rows.append(record)
    return rows


def validate(record: dict, warnings: list[dict]) -> bool:
    qid = record.get("query_id")
    missing = [f for f in REQUIRED if record.get(f) in (None, "")]
    if missing:
        warnings.append({"query_id": qid, "reason": "missing_required", "fields": missing})
        return False
    if record["surface"] not in SURFACE_ENUM:
        warnings.append({"query_id": qid, "reason": "surface_out_of_enum", "surface": record["surface"]})
        return False
    return True


def main(xlsx_path: str = "catalogos_eje4_canal_descubrimiento.xlsx") -> None:
    xlsx = Path(xlsx_path)
    if not xlsx.exists():
        raise SystemExit(f"xlsx not found: {xlsx}")

    now = datetime.now(timezone.utc)
    batch_id = "batch_" + now.strftime("%Y%m%d_%H%M%S")
    out_dir = Path("working/eje4/queries") / batch_id
    out_dir.mkdir(parents=True, exist_ok=True)

    warnings: list[dict] = []
    rows = load_rows(xlsx, warnings)
    by_catalogo: Counter[str] = Counter()
    by_surface: Counter[str] = Counter()
    processed = 0

    for record in rows:
        if not validate(record, warnings):
            continue
        sheet = record.pop("_sheet")
        (out_dir / f"query_{record['query_id']}.json").write_text(
            json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        by_catalogo[sheet] += 1
        by_surface[record["surface"]] += 1
        processed += 1

    manifest = {
        "batch_id": batch_id,
        "generated_at": now.strftime("%Y-%m-%dT%H:%M:%SZ"),
        "xlsx_source": xlsx.name,
        "total_queries": processed,
        "distribution_by_catalogo": {s: by_catalogo.get(s, 0) for s in SHEETS},
        "distribution_by_surface": {s: by_surface.get(s, 0) for s in SURFACE_ENUM},
        "warnings": warnings,
    }
    (out_dir / "batch_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print(f"Batch: {batch_id}")
    print(f"Total queries in xlsx: {len(rows)} | processed: {processed} | warnings: {len(warnings)}")
    print(f"Output: {out_dir}")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "catalogos_eje4_canal_descubrimiento.xlsx")
